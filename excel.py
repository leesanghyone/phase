
import openpyxl,random,datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Side,Border,Alignment
from datetime import datetime 

#뼈대엑셀파일 생성하기.
def create_excel():
    wb= openpyxl.Workbook()
    ws=wb.active

#기존의 뼈대가 되는 정보들이다.
    ws.append(["총작업수","플랫폼","아이디","카톡명","상품명","구매일자","리뷰모음","포토리뷰","포인트금액","작업상태","상품금액","입금일자","입금금액","입금자명"])

#첫번째 행의 디자인을 수정한다.
    border1=Side(color="000000",border_style="thick")
    border_all=Border(left=border1,right=border1,top=border1,bottom=border1)
    alighn=Alignment(horizontal="center",vertical="center")
    ft=Font(size=13,color="700608",bold=True,name="돋움")

    for row in ws.iter_rows(max_row=1):
        for cell in row:
            cell.alignment=alighn
            cell.border=border_all
            cell.font=ft

#모든 행에 넓이 자동조절.
    for col in range(1, 16):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].auto_size = True

#행 높이를 절제한다. 리뷰처럼 길시 높이를 제약을 건다.
    ws.row_dimensions[1].height=30
# ----------------저장하기  --------------------#
    wb.save(f"구매기록지.xlsx")




################################################################
##작업데이터 계속 추가하는 코드다###
def add_exceldata(platform,workcom,kakao,product_title,review,photoreview,pointuse,workstate,product_price):
    now = datetime.now()
#뼈대엑셀파일 생성하기.
    wb= openpyxl.load_workbook(f"구매기록지.xlsx")
    ws=wb.active

#데이터 넣기.
    #총인덱스는,마지막 숫자에 +1을 한다.
    if ws["A"][-1].value=="총작업수":  #ws["A"][-1]첫번째셀값이다. 
        index=1
    else:
        index=int(ws["A"][-1].value)+1

    #기록에 포함되는 데이터.
    # platform="쿠팡"
    # workcom="아무개"
    # kakao="개발중"
    workdate=now.strftime("%Y-%m-%d")
    # review="스타일 적용시, “셀”에 적용한다. 단위를 잘 파악해야 됨. \n 참고)셀 내용이 있어야 스타일이 적용됨."
    # photoreview=" "
    # pointuse= 12000
    # workstate="리뷰작성"
    # product_price=12000
    ws.append([index,platform,workcom,kakao,product_title,workdate,review,photoreview,pointuse,workstate,product_price])

#첫번째 행의 디자인을 수정한다.
    border1=Side(color="000000",border_style="thin")
    border_all=Border(left=border1,right=border1,top=border1,bottom=border1)
    alighn=Alignment(horizontal="center",vertical="center")
    for row in ws.iter_rows(min_row=2): 
        for cell in row:
            cell.alignment=alighn
            cell.border=border_all

#행높이를 절제한다. 리뷰처럼 길시 높이를 제약을 건다.
    ws.row_dimensions[ws.max_row].height=15
    wb.save(f"구매기록지.xlsx")

if __name__ == "__main__":
    try:
        platform="쿠팡"
        workcom="박경희"
        kakao="개발중"
        product_title="경추베개"
        review="너무 좋아요, 우리 모두 그렇게 했으면 좋겠어요. \n 아니야 !!!"
        photoreview="X"
        pointuse=None
        workstate="구매완료"
        product_price=13000
        add_exceldata(platform,workcom,kakao,product_title,review,photoreview,pointuse,workstate,product_price)
    except:
        create_excel()